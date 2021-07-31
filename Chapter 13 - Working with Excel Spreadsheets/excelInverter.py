#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excelInverter.py
Created on Sat Jul 31 18:48:03 2021

@author: paulmason

program that takes an Excel file and for each cell inverts its row and column values

steps:
    1. prompt the user for what will be the Excel file that is inverted and create a blank Workbook obj
    2. use a for loop to get all the row and column values per cell in the given workbook
    3. in the for loop insert the cells value into the new Workbook with the proper inverted position
    4. save the new workbook created as the Excel file given + Inverted.xlsx
"""

#import openpyxl to open workbook given and create and save the inverted workbook object
import openpyxl
#import get_column_letter function to switch from index value to letter values for columns
from openpyxl.utils import get_column_letter

#step 1
#explain the program
print('''This is a program that takes an Excel file and 
creates a new file with the row and column values inverted\n''')
#create a blank excelFile value to prompt the user
excelFile = ''

#to ensure the user enters a proper Excel file use this while loop
while not excelFile.endswith('.xlsx'):
    #prompt the user for an Excel file
    print('Enter an Excel file to invert (with the .xlsx)')
    #save the value given into the excelFile variable
    excelFile = input()

#output that the Excel file is being openned for the user
print('Opening the file: %s...' % excelFile)    
#open the Excel file given
wbGiven = openpyxl.load_workbook(excelFile)
#get the active sheet in the Workbook
sheetGiven = wbGiven.active
#create a blank workbook object to save the inverted values in
wbInverted = openpyxl.Workbook()
#get the active sheet in the blank Workbook
sheetInverted = wbInverted.active

#step 2
#loop through the rows in the sheet given
for row in range(1, sheetGiven.max_row + 1):
    #loop through the columns in the sheet given
    for col in range(1, sheetGiven.max_column + 1):
        #get the column letter from the index of col
        colLetter = get_column_letter(col)
        #output to the user what is happening
        print('Inverting row: %s and column: %s...' % (str(row), colLetter))
        
        #step 3
        #get the inverted column letter from the row value
        invertedColLetter = get_column_letter(row)
        #get the inverted row value from col
        invertedRowVal = col
        #output the inverted row and column, useful debug check and shows user program works
        print('Creating new value at row: %s and column: %s...' % (str(invertedRowVal), invertedColLetter))
        #now set the value for the inverted sheet from the given sheet
        sheetInverted[invertedColLetter + str(invertedRowVal)] = sheetGiven[colLetter + str(row)].value

#step 4
#remove the .xlsx from excelFile variable that was input above
excelFile = excelFile.replace('.xlsx', '')
#add Inverted.xlsx to save a new excel file that is aptly named
excelFile += 'Inverted.xlsx'
#finally save the file
wbInverted.save(excelFile)
#output that the program is done
print('Done!')
        
