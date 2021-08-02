#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excelToTxt.py
Created on Mon Aug  2 10:45:39 2021

@author: paulmason

program that takes the values in each column in an Excel file and inserts them
into a .txt file with each row being its own line and each column being its 
own .txt file

steps:
    1. prompt the user to enter a .xlsx file and use openpyxl to load the wb 
       and get the active sheet
    2. loop through the columns in the sheet, creating a .txt file for each column
    3. loop through the rows in the sheet and write the value of the cell using the 
       write() method and then close the .txt file
"""

#import the openpyxl module to load wb given and get the active sheet
import openpyxl
#import get_column_letter function to switch from index value to letter values for columns
from openpyxl.utils import get_column_letter

#step 1
#explain program to user
print('This program reads an Excel file and creates a .txt file for each column.\n')
#create a variable to store user input
excelFile = ''
#ensure the file given actually ends with the .xlsx extension
while not excelFile.endswith('.xlsx'):
    #prompt the user for an Excel file
    print('Enter a valid Excel file in the CWD (be sure to have the .xlsx):')
    #get the input from the user
    excelFile = input()

#tell user what Excel file is being opened
print('Opening Excel file: %s...' % excelFile)    
#use openpyxl to load in the Excel file given
wb = openpyxl.load_workbook(excelFile)
#get the active sheet of the workbook
sheet = wb.active

#step 2
#loop through the columns of the of the active sheet
for col in range(1, sheet.max_column + 1):
    #get the column letter from the column index
    colLetter = get_column_letter(col)
    #create the filename as 'filetxt' + col .txt
    fName = 'filetxt' + str(col) + '.txt'
    #output to user what text file is being created
    print('Creating text file: %s...' % fName)
    #create the file using open and open it in w+ mode
    file = open(fName, 'w+')
    
    #step 3
    #loop through the rows in the active sheet
    for row in range(1, sheet.max_row + 1):
        #get the cell content from the active sheet
        cellVal = str(sheet[colLetter + str(row)].value)
        #output what line is being written to what file
        print('Writing value at line %s in file %s...' % (str(row), fName))
        #write the cell value plus a newline value to the file
        file.write(cellVal + '\n')
    
    #close the file
    file.close()
    
#output to user that program is done
print('Done!')
        