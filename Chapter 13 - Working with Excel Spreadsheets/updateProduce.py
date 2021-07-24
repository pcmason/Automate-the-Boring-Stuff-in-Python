#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
updateProduce.py
Created on Fri Jul 23 13:21:19 2021

@author: paulmason

program corrects the cost per pound for garlic, celery and lemons because they 
were entered incorrectly

steps:
    1. open the spreadsheet file produceSales.xlsx
    2. for each row check whether the value in column A is celery, lemon or garlic
    3. if it is update the price in column B
    4. save the spreadsheet in a new file (so you do not lose the old spreadsheet)
"""

#import openpyxl to load the workbook and save the changes to a new Workbook
import openpyxl

#use the load_workbook method to open produceSales.xlsx
wb = openpyxl.load_workbook('produceSales.xlsx')
#get the sheet from the Workbook object with all the data, named 'Sheet'
sheet = wb['Sheet']

#create a dictionary with the produce name and the updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

#loop through the rows and update the prices
#skip the first row since its the header
for rowNum in range(2, sheet.max_row):
    #the cell in column 1 (A) stores the produce name
    produceName = sheet.cell(row = rowNum, column = 1).value
    #check if the name returned is in the PRICE_UPDATES dictionary as a key
    if produceName in PRICE_UPDATES:
        #replace the price in column 2 (B) with the value in PRICE_UPDATED
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

#now save the updated Workbook to a new Excel Workbook titled 'updatedProduceSales.xlsx'
wb.save('updatedProduceSales.xlsx')