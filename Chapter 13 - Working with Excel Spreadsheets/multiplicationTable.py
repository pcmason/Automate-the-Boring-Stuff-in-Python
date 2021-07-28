#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
multiplicationTable.py
run program on mac: python3 multiplicationTable.py <N>
Created on Wed Jul 28 00:25:44 2021

@author: paulmason

program that takes the number N from the command line and creates an N X N multiplication table in Excel
row 1 and column A should be used as labels of the table and be in bold

steps:
    1. read in N from the command line, open a blank Workbook object and create row 1 
       and column A in bold with a for loop
    2. use a double for loop to fill in the values for the multiplication table
    3. save the Workbook as 'multiplicationTable.xlsx'
"""

#import openpyxl to open, edit and save a blank Workbook object, use sys to get N from cmd line
#and to exit program when N is not given
import openpyxl, sys
#also import Font from openpyxl to make the labels bold
from openpyxl.styles import Font
#finally import get_column_letter function to switch from index value to letter values
from openpyxl.utils import get_column_letter

#step 1
#ensure there were only 2 arguments in the command line (filename and N)
if len(sys.argv) != 2:
    #if so N was not given or too many arguments given, tell user and close the program
    print("Program was not run properly either with N not given or too many command line arguments!")
    sys.exit()
#else the program was run correctly
else:
    #use a try statement to ensure you can save n as an integer
    try: 
        n = int(sys.argv[1])
    except Exception as err:
        #then n was input incorrectly
        print("Error occured: " + str(err))

#output what is going on to the user
print('Creating \'mutliplication table Excel Workbook...\'')
#now use openpyxl to open a blank Worbook object
wb = openpyxl.Workbook()
#set the sheet to be the only sheet in the new worbook with Active
sheet = wb.active
#and rename the sheet to be title 'N Multiplication Table'
sheet.title = str(n) + ' Mutliplication Table'
#create the label font which is bolded
labelFont = Font(bold = True)

#create the for loop starting at 2 and going up to n + 2
for i in range(2, n+2):
    #set the row value to be i - 1
    sheet['A' + str(i)] = i - 1
    #get the column letter value from the index 
    colLetter = get_column_letter(i)
    #set the column value to be i - 1
    sheet[colLetter + '1'] = i - 1
    #now for both set their font to be labelFont
    sheet['A' + str(i)].font = labelFont
    sheet[colLetter + '1'].font = labelFont

#step 2
#first for loop gets the row values, skipping past 1 as nothing is in A1
for row in range(2, sheet.max_row + 1):
    #print out what row you are on to let the user know what is going on
    print('creating row: ' + str(row) + '...')
    #do the same for loop for the column values
    for col in range(2, sheet.max_column + 1):
        #now get the letter value from the numerical index of the column
        colLetter = get_column_letter(col)
        #now set the Cell value as the multiplication of the values in Cell Ai * Cell colLetter1
        sheet[colLetter + str(row)] = sheet['A' + str(row)].value * sheet[colLetter + '1'].value
        
#step 3
#now everything should be done save the Workbook, I'll title it multiplationTable.xlsx
wb.save('multiplicationTable.xlsx')
#output to the user the program is done and the name of the newly created Excel file
print('Done. Find the multiplication table in file: multiplicationTable.xlsx !')
        
    
        