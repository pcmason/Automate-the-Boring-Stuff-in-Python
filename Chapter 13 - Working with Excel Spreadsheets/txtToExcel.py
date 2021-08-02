#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
txtToExcel.py
Created on Mon Aug  2 00:55:05 2021

@author: paulmason

program that reads the contents of text files in the current directoy and inserts them
into an Excel spreadsheet with one line per row
the lines of file one will be in column A, file two in B, and so on

steps:
    1. get the cd using os library function getcwd()
    2. use openpyxl library to create a new blank Workbook object and get the active sheet
    3. use the os library's listdir() method to find all .txt files
    4. use the readlines() File method to get the list of strings per line in the .txt file then use
       a for loop to insert each line into the blank Workbook object
    5. save the Workbook object and name it 'txtSpreadsheet.xlsx'
"""

#import openpyxl to create a blank workbook object and save to it and os to get the cd and find all txt files
import openpyxl, os
#import get_column_letter function to switch from index value to letter values for columns
from openpyxl.utils import get_column_letter

#step 1
#output what the program does to the user
print('This program puts each line of .txt files in the CWD into an Excel file.')
#get the current working directory
cwd = os.getcwd()
#output the cwd
print('The current working directoy is: %s' % (cwd))

#step 2
#create a blank workbook
wb = openpyxl.Workbook()
#get the active sheet in the workbook
sheet = wb.active

#step 3
#create an empty list to store the text files in
txtFiles = []
#loop through all files in the cwd using listdir() method
for file in os.listdir(cwd):
    #make sure the file ends with '.txt'
    if file.endswith('.txt'):
        #append the txt file to txtFiles
        txtFiles.append(file)

#step 4
#loop through the txtFiles list but get the index to make adding the the workbook easier
for i in range(len(txtFiles)):
    #output to user what file is being openned
    print('Opening file: %s...' % txtFiles[i])
    #open the .txt file
    f = open(txtFiles[i], 'r')
    #store the lines of the file in a list called lines
    lines = f.readlines()

    #loop through lines getting the index of each line to make it easier to add to the workbook
    for j in range(len(lines)):
        #get the column letter from i's value (add one since Excel sheets start at 1 not 0)
        colLetter = get_column_letter(i + 1)
        #output to user what column and row is being created
        print('Inserting text into row %s column %s...' % (str(j+1), colLetter))
        #now insert the value in lines in cell 'ij'
        sheet[colLetter + str(j + 1)] = lines[j]

#step 5
#save the workbook with name txtSpreadsheet
wb.save('txtSpreadsheet.xlsx')
#output to user program is done
print("Done, Excel file created in CWD titled: 'txtSpreadsheet.xlsx'")