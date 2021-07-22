#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
readCensusPop.py
Created on Tue Jul 20 23:06:34 2021

@author: paulmason

program that reads from the Excel file censuspopdata.xlsx: a file that has census data from 2010
counts the total number of census tracts for each county and the total population of each county
the file created will be census2010.py which can be imported to get the census data

steps:
    1. open and read the cells of censuspopdata.xlsx using openpyxl module
    2. calculate all the tract and population data and store it in a data structure
    3. write the data structure to a text file with the .py extension using the pprint module
"""

#import openpyxl to load the Excel sheet and pprint to write the data at the end to a .py file
import openpyxl, pprint

#step 1
#output that the Excel document is being opened
print('Opening workbook...')
#use load_workbook to open censuspopdata Excel file
wb = openpyxl.load_workbook('censuspopdata.xlsx')
#get the sheet with the census data named 'Population by Census Tract'
sheet = wb['Population by Census Tract']
#create an empty dict to store the number of tracts for each county and its population
countyData = {}

#step 2
#fill in countyData with each country's population and tract
#output that rows are being read to the user
print('Reading rows...')

#begin iterating over the rows, beginning at 2 to skip the title cells
for row in range(2, sheet.max_row + 1):
    #each row has data for one census tract
    #the B column holds what state it is
    state = sheet['B' + str(row)].value
    #the C column holds the county in the state
    county = sheet['C' + str(row)].value
    #the D column holds the population data
    pop = sheet['D' + str(row)].value
    
    #ensure the key for the state exists and if it does not create a key for the state with setdefault
    countyData.setdefault(state,{})
    #ensure the key for the county exists and if not use setdefault to create one
    countyData[state].setdefault(county,{'tracts': 0, 'pop': 0})
    
    #each row represents one census tract so increment by 1
    countyData[state][county]['tracts'] += 1
    #increase the county population by pop
    countyData[state][county]['pop'] += int(pop)
    
#step 3
#open a new text file and write the contents of countryData to it
#output the program has moved onto the writing to file stage
print('Writing results...')
#open a file for results naming it census2010.py and enabling writing to it
resultFile = open('census2010.py', 'w')
#use the pformat function to make this valid python code so you can import census2010 and use it
resultFile.write('allData =' +pprint.pformat(countyData))
#close the file
resultFile.close()
#print the program is done running
print('Done')