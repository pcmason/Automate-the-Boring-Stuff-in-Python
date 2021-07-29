#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
blankRowInserter.py
how to run on mac: python3 blankRowInserter.py <EXCEL_FILE> <N> <M>
python3 blankRowInserter.py produceSales.xlsx 3 2
Created on Thu Jul 29 01:50:02 2021

@author: paulmason

program that takes an Excel file and two numbers (N & M) from the cmd line and
starting at row N inserts M blank rows into the spreadsheet

steps:
    1. ensure that the cmd line arguments were entered correctly
    2. open the Excel file given and loop through adding the blank rows where needed
       and moving all other rows down
    3. save the changes to a new Excel files adding BlankRows to the filename
"""
#import sys to check cmd line arguments and openpyxl to open and save changes
#to the Excel file given
import openpyxl, sys
#import get_column_letter function to switch from index value to letter values for columns
from openpyxl.utils import get_column_letter

#step 1
#ensure there are only 4 cmd line arguments
if len(sys.argv) == 4:
    #then first ensure the second argument ends with .xlsx
    if not sys.argv[1].endswith('.xlsx'):
        #tell the user they entered the Excel file incorrectly and exit
        print('Incorrectly input the Excel file!')
        sys.exit()
    else:
        #then set the excelFile variable to the second argument
        excelFile = sys.argv[1]
    #then set N and M
    n = int(sys.argv[2])
    m = int(sys.argv[3])
    print('N is %s and M is %s' % (str(n),str(m)))
else:
    #either too few or too many arguments so inform user and exit program
    print('Too many or too few command line arguments!')
    sys.exit()
    
#step 2
#output to user what is happening
print('Opening Excel file: %s...' % (excelFile))
#create the workbook object using the load_workbook method
wb = openpyxl.load_workbook(excelFile)
#get the active sheet in the workbook
sheet = wb.active

#ensure that N is not greater than the max_row in the sheet
if n > (sheet.max_row + 1):
    #tell user that they put in too large of a row to begin at and exit the program
    print('Excel file: %s does not have enough rows to insert blank rows!' % (excelFile))
    sys.exit()

#before inserting the blank rows, loop through the excel sheet starting
#at the end of the rows in the Excel file and going backwards to n
for row in range(sheet.max_row + 1, n - 1, -1):
    #update to user what is happening
    print('Moving row %s...' % (str(row)))
    #loop through column values like normal
    for col in range(1, sheet.max_column + 1):
        #get the column letter from the index
        colLetter = get_column_letter(col)
        #now move all row values down by m
        sheet[colLetter + str(row + m)] = sheet[colLetter + str(row)].value

#then insert the blank rows beginning at N going to the end of the file
for row in range(n, n + m):
    #output to user what is happening
    print('Inserting blank row at row: ' + str(row) + '...')
    #loop through every column value
    for col in range(1, sheet.max_column + 1):
        #get the column letter from the column index
        colLetter = get_column_letter(col)
        #create blank value at row and colLetter given
        sheet[colLetter + str(row)] = ''
         
    

#step 3
#create a new file name for the new sheet with blank rows inserted
#first remove the .xlsx from excelFile
excelFile = excelFile.replace('.xlsx', '')
#now create the new filename by adding BlankRows.xlsx
excelFile += 'BlankRows.xlsx'
#now save the new excelFile
wb.save(excelFile)
#output that the program is done
print('Done!')
    

