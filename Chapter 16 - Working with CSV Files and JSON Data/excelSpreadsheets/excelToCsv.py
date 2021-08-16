#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excelToCsv.py
Created on Sun Aug 15 19:53:49 2021

@author: paulmason

program that converts all Excel files in the current directory to CSV file
has to create a CSV file for each sheet in the Excel file

steps:
    1. loop through all the files in the cwd, only use the Excel files
    2. loop through all the sheets in each Excel file
    3. create a CSV writer object for each sheet and add the contents of the Excel sheet to it
    4. save the CSV file and close it
"""

#import csv to create writer objects and os to get the files in the cwd and openpyxl to work with Excel files
import csv, os, openpyxl
#import the method from openpyxl to convert column numbers into letters
from openpyxl.utils import get_column_letter

#step 1 - loop through the Excel files in the cwd
#go through every file in the cwd
for excelFile in os.listdir('.'):
    #skip non-xlsx files, load the workbook object
    if not excelFile.endswith('.xlsx'):
        continue
    
    #load the workbook object
    wb = openpyxl.load_workbook(excelFile)
    
    #step 2 - loop through all the sheets in each Excel file
    for sheetName in wb.get_sheet_names():
        #loop through every sheet in the workbook
        sheet = wb.get_sheet_by_name(sheetName)

        #create the CSV filename from the Excel filename and sheet title = <excel_fn>_<sheetname>.csv
        csvFilename = excelFile.replace('.xlsx', '_' + sheet.title + '.csv')
        print('Creating CSV file %s...' % csvFilename)
        #open/create the new CSV file
        csvFile = open(csvFilename, 'w', newline = '')
        
        #step 3 - create csv writer object and add all of the cells from the Excel sheet to it
        #create the csv.writer object for this CSV file
        csvWriter = csv.writer(csvFile)

        #loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            #append each cell to this list
            rowData = []    
            
            #loop through each cell in the row
            for colNum in range(1, sheet.max_column + 1):
                #convert the column number to the appropriate column index
                colLetter = get_column_letter(colNum)
                #append each cell's data to rowData
                rowData.append(sheet[colLetter + str(rowNum)].value)

            #write the rowData list to the CSV file
            csvWriter.writerow(rowData)
            
        #step 4 - save and close the CSV file
        #close the CSV file
        csvFile.close()

#once done with the above loop the program will be done
print('Done!')        
